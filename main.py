from colors import *
from tkinter import *
from tkinter import messagebox
import random
import string
import os
import webbrowser
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pyperclip

user_data_file_name = "Users.xlsx"
log_data_file_name = "Log_Data.xlsx"

currentDirectory = os.getcwd()

user_data_file_path = os.path.join(currentDirectory, user_data_file_name)
log_data_file_path = os.path.join(currentDirectory, log_data_file_name)

# Create a new workbook if it doesn't exist
def create_workbook_if_not_exists(file_path, sheet_name, headers):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
        wb.save(file_path)
    else:
        wb = load_workbook(file_path)
        ws = wb.active
    return wb, ws

user_headers = ["ID", "First Name", "Middle Name", "Last Name"]
log_headers = ["ID", "Student Name", "Time", "Date"]

user_wb, user_ws = create_workbook_if_not_exists(user_data_file_path, "User Data", user_headers)
log_wb, log_ws = create_workbook_if_not_exists(log_data_file_path, "Log Data", log_headers)

window = Tk()
window.title("Attendance System")
window.geometry("650x350")

Home = Frame(window, bg=colorLightestBlue)
Register = Frame(window, bg=colorLightestBlue)

def handle_log_time():
    student_id = inp.get()
    
    if not student_id:  # Check if the input is empty
        messagebox.showerror("Error", "Student ID cannot be empty!")
        return

    now = datetime.now()
    current_time = now.strftime("%I:%M:%S %p")
    current_date = now.strftime("%Y-%m-%d")
    student_name = None
    
    for row in user_ws.iter_rows(min_col=1, max_col=4, values_only=True):
        if row[0] == student_id:
            student_name = f"{row[3]}, {row[1]} {row[2]}"
            break

    if student_name:
        log_ws.append([student_id, student_name, current_time, current_date])
        log_wb.save(log_data_file_name)
        messagebox.showinfo("Success", f"Successfully logged time for {student_name}")
        inp.delete(0, END)
    else:
        messagebox.showerror("Error", f"Student ID {student_id} not found!")

def handle_key_press(event: object):
    if event.keysym == "Return":
        handle_log_time()

def goto_page(page):
    Home.pack_forget()
    Register.pack_forget()

    if page == "home":
        Home.pack(fill="both", expand=True)
    elif page == "register":
        Register.pack(fill="both", expand=True)

def generate_random_id():
    part1 = ''.join(random.choice(string.digits) for _ in range(4))
    part2 = ''.join(random.choice(string.digits) for _ in range(4))
    random_id = f"{part1}-{part2}"
    return random_id

def check_duplicate_id(ws, random_id):
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        if row[0] == random_id:
            return True
    return False

def handle_register_student(first, middle, last):
    if not first or not middle or not last:  # Check if any field is empty
        messagebox.showerror("Error", "All fields must be filled!")
        return

    random_id = generate_random_id()
    while check_duplicate_id(user_ws, random_id):
        random_id = generate_random_id()
    user_ws.append([random_id, first, middle, last])
    user_wb.save(user_data_file_name)
    
    messagebox.showinfo("Registration Successful", f"Your ID is: {random_id}, which has been automatically copied to the clipboard. You can now paste it in the log data.")
    pyperclip.copy(random_id)
    
    first_name.delete(0, END)
    middle_name.delete(0, END)
    last_name.delete(0, END)

    print(f"Registered Student: {first} {middle} {last} with ID {random_id}")
    return random_id

# Home Page -------------------------------------------------

Label(
    Home,
    font=("Arial", 14),
    text = "Attendance System by Khian Victory D. Calderon",
    bg=colorLightestBlue,
    fg=colroDarkGray
).pack(padx = 5, pady = 10)

Label(
    Home,
    text="Enter student ID: ",
    font=("Arial", 20),
    bg=colorLightestBlue,
    fg=colroDarkGray
).pack(padx=5, pady=5)

inp = Entry(Home, font=("Arial", 20))
inp.pack(padx=5, pady=5)
inp.bind("<Return>", handle_key_press)

Button(
    Home,
    text="Log Time",
    bg=colorLighterBlue,
    fg=colorLightWhite,
    activeforeground=colorLightWhite,
    activebackground=colorLightBlue,
    command=handle_log_time,
    font=("Arial", 20)
).pack(padx=5, pady=5)

Button(
    Home,
    text="Register Student",
    bg=colorLighterBlue,
    fg=colorLightWhite,
    activeforeground=colorLightWhite,
    activebackground=colorLightBlue,
    command=lambda: goto_page("register"),
    font=("Arial", 20)
).pack(padx=5, pady=5)

Button(
    Home,
    text="Khian's Website",
    bg=colorLighterBlue,
    fg=colorLightWhite,
    activeforeground=colorLightWhite,
    activebackground=colorLightBlue,
    command=lambda: webbrowser.open("https://khian.netlify.app/"),
    font=("Arial", 14)
).pack(padx=5, pady=5)

# End Home Page -------------------------------------------------

# Register Page -------------------------------------------------

Register.columnconfigure(0, weight=1)
Register.columnconfigure(1, weight=2)

# First name
Label(Register, text="First Name: ", font=("Arial", 20), bg=colorLightestBlue).grid(row=0, column=0, padx=5, pady=5)
first_name = Entry(Register, font=("Arial", 20))
first_name.grid(row=0, column=1, padx=5, pady=5, sticky=NSEW)

# Middle name
Label(Register, text="Middle Name: ", font=("Arial", 20), bg=colorLightestBlue).grid(row=1, column=0, padx=5, pady=5)
middle_name = Entry(Register, font=("Arial", 20))
middle_name.grid(row=1, column=1, padx=5, pady=5, sticky=NSEW)

# Last name
Label(Register, text="Last Name: ", font=("Arial", 20), bg=colorLightestBlue).grid(row=2, column=0, padx=5, pady=5)
last_name = Entry(Register, font=("Arial", 20))
last_name.grid(row=2, column=1, padx=5, pady=5, sticky=NSEW)

Button(
    Register,
    text="Register",
    bg=colorLighterBlue,
    fg=colorLightWhite,
    activeforeground=colorLightWhite,
    activebackground=colorLightBlue,
    command=lambda: handle_register_student(first_name.get(), middle_name.get(), last_name.get()),
    font=("Arial", 20)
).grid(row=3, column=0, columnspan=2, pady=20)

Button(
    Register,
    text="Home",
    bg=colorLighterBlue,
    fg=colorLightWhite,
    activeforeground=colorLightWhite,
    activebackground=colorLightBlue,
    command=lambda: goto_page("home"),
    font=("Arial", 20)
).grid(row=4, column=0, columnspan=2, pady=20)

# End Register Page -------------------------------------------------

Home.pack(fill="both", expand=True)

window.mainloop()
